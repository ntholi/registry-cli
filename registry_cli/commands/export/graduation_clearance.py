import os
from datetime import datetime
from typing import List, Optional

import click
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, func
from sqlalchemy.orm import Session

from registry_cli.models import (
    Clearance,
    GraduationClearance,
    GraduationRequest,
    PaymentReceipt,
    Program,
    School,
    Structure,
    Student,
    StudentProgram,
)


def export_approved_graduation_students(db: Session) -> None:
    """
    Export students who have been approved for graduation requests clearance by all required departments.

    This command exports students who have:
    1. A graduation request
    2. Approved clearances from finance, library, and academic departments
    3. Their information includes: student number, student names, faculty, program, and receipt numbers

    The exported file includes: student number, name, faculty (school), program name, and receipt numbers.
    """
    click.echo(
        "Finding students with approved graduation clearances from finance, library, and academic departments..."
    )

    # Define the three required departments
    required_departments = ["finance", "library", "academic"]
    required_dept_count = len(required_departments)

    click.echo(f"Required departments: {required_departments}")

    # Single optimized query to get all approved students with their details and payment receipts
    query = (
        db.query(
            GraduationRequest.id.label("graduation_request_id"),
            StudentProgram.std_no.label("student_number"),
            Student.name.label("student_name"),
            School.name.label("faculty"),
            Program.name.label("program_name"),
            PaymentReceipt.receipt_no.label("receipt_no"),
            PaymentReceipt.payment_type.label("payment_type"),
        )
        .select_from(GraduationRequest)
        .join(StudentProgram, GraduationRequest.student_program_id == StudentProgram.id)
        .join(Student, StudentProgram.std_no == Student.std_no)
        .join(Structure, StudentProgram.structure_id == Structure.id)
        .join(Program, Structure.program_id == Program.id)
        .join(School, Program.school_id == School.id)
        .join(
            GraduationClearance,
            GraduationRequest.id == GraduationClearance.graduation_request_id,
        )
        .join(Clearance, GraduationClearance.clearance_id == Clearance.id)
        .outerjoin(
            PaymentReceipt, PaymentReceipt.graduation_request_id == GraduationRequest.id
        )
        .filter(
            and_(
                Clearance.status == "approved",
                Clearance.department.in_(required_departments),
            )
        )
        .group_by(
            GraduationRequest.id,
            StudentProgram.std_no,
            Student.name,
            School.name,
            Program.name,
            PaymentReceipt.receipt_no,
            PaymentReceipt.payment_type,
        )
        .having(func.count(func.distinct(Clearance.department)) == required_dept_count)
        .order_by(GraduationRequest.id, PaymentReceipt.receipt_no)
    )

    results = query.all()

    if not results:
        click.secho(
            "No students found with approved graduation clearances from all required departments.",
            fg="yellow",
        )
        return

    # Group results by graduation request to organize payment receipts
    student_data = {}
    for row in results:
        grad_req_id = row.graduation_request_id

        if grad_req_id not in student_data:
            student_data[grad_req_id] = {
                "graduation_request_id": grad_req_id,
                "student_number": row.student_number,
                "student_name": row.student_name,
                "faculty": row.faculty,
                "program_name": row.program_name,
                "graduation_fee_receipts": [],
                "graduation_gown_receipts": [],
                "all_receipts": [],
            }

        # Add payment receipt if it exists
        if row.receipt_no:
            if row.payment_type == "graduation_fee":
                student_data[grad_req_id]["graduation_fee_receipts"].append(
                    row.receipt_no
                )
            elif row.payment_type == "graduation_gown":
                student_data[grad_req_id]["graduation_gown_receipts"].append(
                    row.receipt_no
                )
            student_data[grad_req_id]["all_receipts"].append(row.receipt_no)

    # Convert to list and format receipt numbers
    approved_students = []
    for student in student_data.values():
        # Remove duplicates and format receipt strings
        student["graduation_fee_receipts"] = list(
            set(student["graduation_fee_receipts"])
        )
        student["graduation_gown_receipts"] = list(
            set(student["graduation_gown_receipts"])
        )
        student["all_receipts"] = list(set(student["all_receipts"]))

        approved_students.append(
            {
                "graduation_request_id": student["graduation_request_id"],
                "student_number": student["student_number"],
                "student_name": student["student_name"],
                "faculty": student["faculty"],
                "program_name": student["program_name"],
                "graduation_fee_receipts": (
                    ", ".join(student["graduation_fee_receipts"])
                    if student["graduation_fee_receipts"]
                    else "N/A"
                ),
                "graduation_gown_receipts": (
                    ", ".join(student["graduation_gown_receipts"])
                    if student["graduation_gown_receipts"]
                    else "N/A"
                ),
                "all_receipts": (
                    ", ".join(student["all_receipts"])
                    if student["all_receipts"]
                    else "N/A"
                ),
            }
        )

    if not approved_students:
        click.secho("No valid students found after processing.", fg="yellow")
        return

    click.echo(
        f"Found {len(approved_students)} students with full departmental approval"
    )

    # Sort students by graduation request ID
    approved_students.sort(key=lambda x: x["graduation_request_id"])

    # Export to Excel
    output_dir = "exports"
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"approved_graduation_clearance_{timestamp}.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        ws = wb.create_sheet("Approved Graduation Students")
    else:
        ws.title = "Approved Graduation Students"

    # Set up headers
    headers = [
        "Student Number",
        "Student Name",
        "Faculty",
        "Program",
        "Graduation Fee Receipts",
        "Graduation Gown Receipts",
        "All Receipt Numbers",
        "Graduation Request ID",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

    # Add data rows
    for row, student in enumerate(approved_students, 2):
        ws.cell(row=row, column=1, value=student["student_number"])
        ws.cell(row=row, column=2, value=student["student_name"])
        ws.cell(row=row, column=3, value=student["faculty"])
        ws.cell(row=row, column=4, value=student["program_name"])
        ws.cell(row=row, column=5, value=student["graduation_fee_receipts"])
        ws.cell(row=row, column=6, value=student["graduation_gown_receipts"])
        ws.cell(row=row, column=7, value=student["all_receipts"])
        ws.cell(row=row, column=8, value=student["graduation_request_id"])

    # Auto-size columns
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save the file
    wb.save(excel_path)

    click.secho(
        f"Successfully exported approved graduation students to: {excel_path}",
        fg="green",
    )

    # Display summary
    click.echo(f"\nSummary:")
    click.echo(
        f"- Total students with full departmental approval: {len(approved_students)}"
    )

    # Show breakdown by faculty
    from collections import Counter

    faculty_counts = Counter(student["faculty"] for student in approved_students)
    click.echo(f"\nFaculty breakdown:")
    for faculty, count in faculty_counts.most_common():
        click.echo(f"- {faculty}: {count} students")

    # Show breakdown by program
    program_counts = Counter(student["program_name"] for student in approved_students)
    click.echo(f"\nProgram breakdown:")
    for program, count in program_counts.most_common():
        click.echo(f"- {program}: {count} students")

    # Show payment receipt statistics
    students_with_fee_receipts = sum(
        1 for s in approved_students if s["graduation_fee_receipts"] != "N/A"
    )
    students_with_gown_receipts = sum(
        1 for s in approved_students if s["graduation_gown_receipts"] != "N/A"
    )
    students_with_any_receipts = sum(
        1 for s in approved_students if s["all_receipts"] != "N/A"
    )

    click.echo(f"\nPayment receipt statistics:")
    click.echo(f"- Students with graduation fee receipts: {students_with_fee_receipts}")
    click.echo(
        f"- Students with graduation gown receipts: {students_with_gown_receipts}"
    )
    click.echo(f"- Students with any payment receipts: {students_with_any_receipts}")
    click.echo(
        f"- Students without payment receipts: {len(approved_students) - students_with_any_receipts}"
    )

import os
from collections import defaultdict
from datetime import datetime
from typing import Dict, List

import click
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from registry_cli.models import (
    Program,
    RegistrationRequest,
    School,
    Structure,
    Student,
    StudentProgram,
)


def format_semester(semester_number: int) -> str:
    """Format semester number to YnSm format (e.g., 1 -> Y1S1, 2 -> Y1S2, 3 -> Y2S1)."""
    if not semester_number or semester_number < 1:
        return f"Y1S1"

    year = ((semester_number - 1) // 2) + 1
    semester = ((semester_number - 1) % 2) + 1
    return f"Y{year}S{semester}"


def export_students_by_school(db: Session) -> None:
    """Export registered students grouped by school to Excel with separate sheets."""
    # Query to get all registered students with their school information
    # Joins: RegistrationRequest -> Student -> StudentProgram -> Structure -> Program -> School
    query = (
        db.query(
            School.code.label("school_code"),
            Student.name.label("student_name"),
            Student.std_no,
            RegistrationRequest.semester_number,
            Program.name.label("program_name"),
        )
        .join(Student, RegistrationRequest.std_no == Student.std_no)
        .join(StudentProgram, Student.std_no == StudentProgram.std_no)
        .join(Structure, StudentProgram.structure_id == Structure.id)
        .join(Program, Structure.program_id == Program.id)
        .join(School, Program.school_id == School.id)
        .filter(RegistrationRequest.status == "registered")
        .filter(StudentProgram.status == "Active")
        .distinct()
        .order_by(School.code, Program.name, RegistrationRequest.semester_number)
    )

    results = query.all()

    if not results:
        click.secho("No registered students found.", fg="yellow")
        return

    school_data: Dict[str, Dict[int, Dict]] = defaultdict(dict)

    for school_code, student_name, std_no, semester_number, program_name in results:
        school_data[school_code][std_no] = {
            "name": student_name,
            "student_number": std_no,
            "semester": format_semester(semester_number),
            "program": program_name,
            "semester_order": semester_number,
        }

    output_dir = "exports"
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"students_by_school_{timestamp}.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)

    wb = Workbook()
    if wb.active:
        wb.remove(wb.active)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    total_students = 0

    for school_code in sorted(school_data.keys()):
        students_dict = school_data[school_code]
        students = list(students_dict.values())

        # Sort students by program then by semester number
        students.sort(key=lambda x: (x["program"], x["semester_order"]))

        total_students += len(students)

        safe_sheet_name = school_code[:31]
        if safe_sheet_name in [ws.title for ws in wb.worksheets]:
            safe_sheet_name = f"{safe_sheet_name[:28]}_{len(wb.worksheets)}"

        ws = wb.create_sheet(title=safe_sheet_name)

        headers = ["Student Name", "Student Number", "Program", "Semester"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill

        for row, student in enumerate(students, 2):
            ws.cell(row=row, column=1, value=student["name"])
            ws.cell(row=row, column=2, value=student["student_number"])
            ws.cell(row=row, column=3, value=student["program"])
            ws.cell(row=row, column=4, value=student["semester"])

        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].auto_size = True
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)

    click.secho(f"Successfully exported student data to: {excel_path}", fg="green")

    click.echo(f"\nSummary:")
    click.echo(f"- Total schools: {len(school_data)}")
    click.echo(f"- Total registered students: {total_students}")

    click.echo(f"\nSchool breakdown:")
    for school_name in sorted(school_data.keys()):
        student_count = len(school_data[school_name])
        click.echo(f"- {school_name}: {student_count} students")

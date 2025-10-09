import os
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

import click
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy import and_, or_
from sqlalchemy.orm import Session, joinedload

from registry_cli.commands.approve.academic_graduation import (
    get_outstanding_from_structure,
    get_student_programs,
)
from registry_cli.grade_definitions import calculate_cgpa_from_semesters
from registry_cli.models import (
    Clearance,
    GraduationClearance,
    GraduationRequest,
    Program,
    School,
    SemesterModule,
    Structure,
    Student,
    StudentModule,
    StudentProgram,
    StudentSemester,
)


def has_no_pending_issues(db: Session, std_no: int) -> bool:
    """
    Check if a student has no pending academic issues using the same logic as approve_academic_graduation.

    Returns True if the student has no failed never repeated modules and no never attempted modules.
    """
    try:
        programs = get_student_programs(db, std_no)
        if not programs:
            return False

        outstanding = get_outstanding_from_structure(db, programs)

        # No pending issues if both lists are empty
        return (
            len(outstanding["failedNeverRepeated"]) == 0
            and len(outstanding["neverAttempted"]) == 0
        )
    except Exception as e:
        click.echo(f"Error checking pending issues for student {std_no}: {str(e)}")
        return False


def get_pending_issues_details(db: Session, std_no: int) -> Dict[str, List[Dict]]:
    """
    Get detailed pending issues for a student.
    Returns the same structure as get_outstanding_from_structure.
    """
    try:
        programs = get_student_programs(db, std_no)
        if not programs:
            return {"failedNeverRepeated": [], "neverAttempted": []}

        outstanding = get_outstanding_from_structure(db, programs)
        return outstanding
    except Exception as e:
        click.echo(f"Error getting pending issues for student {std_no}: {str(e)}")
        return {"failedNeverRepeated": [], "neverAttempted": []}


def get_students_expected_to_graduate(
    db: Session,
    graduation_year: int,
    completion_terms: List[str],
    program_levels: List[str],
) -> List[Dict]:
    """
    Get all students expected to graduate based on:
    1. Registration date: comparing reg_date year with graduation year
       - Certificate: 1 year duration
       - Diploma: 3 years duration
       - Degree: 4 years duration
    2. Completion terms: students with specified terms who completed required semesters
       - Certificate: must have semesters 1-2
       - Diploma: must have semesters 1-6
       - Degree: must have semesters 1-8 OR semesters 6-8 (for 3-year programs)

    Only students in the specified program_levels are included.

    Returns list of student dictionaries with basic info.
    """
    expected_students_map = {}  # Use dict to avoid duplicates by std_no

    # Method 1: Based on registration date
    click.echo(
        f"Finding students expected to graduate in {graduation_year} based on registration date..."
    )

    expected_years = {"certificate": 1, "diploma": 3, "degree": 4}

    for level, years in expected_years.items():
        # Skip levels not in the specified program_levels
        if level not in program_levels:
            continue

        target_year = graduation_year - years

        students_query = (
            db.query(
                StudentProgram.std_no,
                StudentProgram.reg_date,
                Program.level.label("program_level"),
                Program.name.label("program_name"),
                School.name.label("school_name"),
                Student.name.label("student_name"),
            )
            .join(Structure, StudentProgram.structure_id == Structure.id)
            .join(Program, Structure.program_id == Program.id)
            .join(School, Program.school_id == School.id)
            .join(Student, StudentProgram.std_no == Student.std_no)
            .filter(
                and_(
                    StudentProgram.status.in_(["Active", "Completed"]),
                    Program.level == level,
                    StudentProgram.reg_date.isnot(None),
                )
            )
            .all()
        )

        # Filter by year only (not full date)
        for student_data in students_query:
            if student_data.reg_date:
                reg_year = int(student_data.reg_date.split("-")[0])
                if reg_year == target_year:
                    expected_students_map[student_data.std_no] = {
                        "std_no": student_data.std_no,
                        "student_name": student_data.student_name,
                        "school_name": student_data.school_name,
                        "program_name": student_data.program_name,
                        "program_level": student_data.program_level,
                        "criteria": f"Reg date year {target_year} ({level}, {years} years)",
                    }

        click.echo(
            f"Found {sum(1 for s in expected_students_map.values() if s['program_level'] == level)} "
            f"{level} students with reg_date year {target_year}"
        )

    # Method 2: Based on completion terms
    click.echo(
        f"Finding students with completion terms: {', '.join(completion_terms)}..."
    )

    # Define semester requirements by level
    semester_requirements = {
        "certificate": {1, 2},
        "diploma": {1, 2, 3, 4, 5, 6},
        "degree": [{1, 2, 3, 4, 5, 6, 7, 8}, {6, 7, 8}],  # Either full 8 or last 3
    }

    # Get all students with the specified terms and program levels
    students_with_terms = (
        db.query(
            StudentProgram.std_no,
            StudentProgram.id.label("program_id"),
            Program.level.label("program_level"),
            Program.name.label("program_name"),
            School.name.label("school_name"),
            Student.name.label("student_name"),
        )
        .join(StudentSemester, StudentProgram.id == StudentSemester.student_program_id)
        .join(Structure, StudentProgram.structure_id == Structure.id)
        .join(Program, Structure.program_id == Program.id)
        .join(School, Program.school_id == School.id)
        .join(Student, StudentProgram.std_no == Student.std_no)
        .filter(
            and_(
                StudentProgram.status.in_(["Active", "Completed"]),
                StudentSemester.term.in_(completion_terms),
                Program.level.in_(program_levels),  # Filter by specified program levels
            )
        )
        .distinct()
        .all()
    )

    click.echo(
        f"Found {len(students_with_terms)} students with specified completion terms and program levels"
    )

    # Check each student's semester completion
    for student_data in students_with_terms:
        level = student_data.program_level

        # Skip if level is not in the specified program_levels (extra safety check)
        if level not in program_levels:
            continue

        # Get all semesters for this student's program
        semesters = (
            db.query(StudentSemester.semester_number)
            .filter(StudentSemester.student_program_id == student_data.program_id)
            .filter(
                StudentSemester.status.notin_(
                    ["Deleted", "Deferred", "DroppedOut", "Withdrawn"]
                )
            )
            .all()
        )

        semester_numbers = {s.semester_number for s in semesters if s.semester_number}

        # Check if student meets semester requirements
        meets_requirements = False
        if level in ["certificate", "diploma"]:
            required_semesters = semester_requirements[level]
            meets_requirements = required_semesters.issubset(semester_numbers)
        elif level == "degree":
            # Check either full 8 semesters or last 3 semesters (6-8)
            option1, option2 = semester_requirements["degree"]
            meets_requirements = option1.issubset(semester_numbers) or option2.issubset(
                semester_numbers
            )

        if meets_requirements:
            expected_students_map[student_data.std_no] = {
                "std_no": student_data.std_no,
                "student_name": student_data.student_name,
                "school_name": student_data.school_name,
                "program_name": student_data.program_name,
                "program_level": student_data.program_level,
                "criteria": f"Completion terms {', '.join(completion_terms)}",
            }

    click.echo(f"Total expected students: {len(expected_students_map)}")
    return list(expected_students_map.values())


def get_non_graduating_students(
    db: Session, expected_students: List[Dict]
) -> List[Dict]:
    """
    From the list of expected students, identify those with pending academic issues.
    These are students who should graduate but have:
    - Failed modules that were never repeated
    - Required modules that were never attempted
    """
    non_graduating_students = []

    click.echo("Checking expected students for pending academic issues...")

    for index, student in enumerate(expected_students, 1):
        if index % 50 == 0:
            click.echo(f"Checked {index}/{len(expected_students)} students...")

        std_no = student["std_no"]

        # Get detailed pending issues once to avoid redundant computations
        pending_issues = get_pending_issues_details(db, std_no)

        # Only include if they actually have pending issues
        failed_never_repeated = pending_issues.get("failedNeverRepeated", [])
        never_attempted = pending_issues.get("neverAttempted", [])

        if not failed_never_repeated and not never_attempted:
            continue

        non_graduating_students.append(
            {
                "student_number": std_no,
                "student_name": student["student_name"],
                "school_name": student["school_name"],
                "program_name": student["program_name"],
                "program_level": student["program_level"],
                "criteria": student.get("criteria", ""),
                "failed_never_repeated": failed_never_repeated,
                "never_attempted": never_attempted,
            }
        )

    click.echo(
        f"Found {len(non_graduating_students)} expected students with pending issues"
    )
    return non_graduating_students


def calculate_graduation_statistics(
    db: Session, graduating_students: List[Dict], non_graduating_students: List[Dict]
) -> Dict:
    """
    Calculate graduation statistics by school and program, similar to breakdown sheet format.
    """
    from collections import defaultdict

    # Structure: stats[school_name][program_name] = {graduating, non_graduating, expected, percentage}
    school_program_stats = defaultdict(
        lambda: defaultdict(
            lambda: {
                "graduating": 0,
                "non_graduating": 0,
                "expected": 0,
                "percentage": 0.0,
            }
        )
    )
    school_totals = defaultdict(
        lambda: {"graduating": 0, "non_graduating": 0, "expected": 0, "percentage": 0.0}
    )
    overall_stats = {
        "graduating": 0,
        "non_graduating": 0,
        "expected": 0,
        "percentage": 0.0,
    }

    # Count graduating students by school and program
    for student in graduating_students:
        school = student["school_name"]
        program = student["program_name"]

        school_program_stats[school][program]["graduating"] += 1
        school_totals[school]["graduating"] += 1
        overall_stats["graduating"] += 1

    # Count non-graduating students by school and program
    for student in non_graduating_students:
        school = student["school_name"]
        program = student["program_name"]

        school_program_stats[school][program]["non_graduating"] += 1
        school_totals[school]["non_graduating"] += 1
        overall_stats["non_graduating"] += 1

    # Calculate totals and percentages for each program
    for school in school_program_stats:
        for program in school_program_stats[school]:
            stats = school_program_stats[school][program]
            stats["expected"] = stats["graduating"] + stats["non_graduating"]
            if stats["expected"] > 0:
                stats["percentage"] = (stats["graduating"] / stats["expected"]) * 100
            else:
                stats["percentage"] = 0.0

    # Calculate totals and percentages for each school
    for school in school_totals:
        stats = school_totals[school]
        stats["expected"] = stats["graduating"] + stats["non_graduating"]
        if stats["expected"] > 0:
            stats["percentage"] = (stats["graduating"] / stats["expected"]) * 100
        else:
            stats["percentage"] = 0.0

    # Calculate overall percentage
    overall_stats["expected"] = (
        overall_stats["graduating"] + overall_stats["non_graduating"]
    )
    if overall_stats["expected"] > 0:
        overall_stats["percentage"] = (
            overall_stats["graduating"] / overall_stats["expected"]
        ) * 100
    else:
        overall_stats["percentage"] = 0.0

    return {
        "school_program_stats": dict(school_program_stats),
        "school_totals": dict(school_totals),
        "overall_stats": overall_stats,
    }


def calculate_cgpa_and_classification_for_program(
    db: Session, std_no: int, program: StudentProgram
) -> Tuple[Optional[float], str]:
    """
    Calculate CGPA and determine classification for a student based on a specific program.
    Uses the same logic as the JavaScript implementation.

    Returns:
        Tuple of (CGPA, Classification)
    """
    try:
        if not program:
            return None, "No Program Provided"

        # Get all semesters for the specified program (excluding deleted/deferred/etc)
        semesters = (
            db.query(StudentSemester)
            .filter(StudentSemester.student_program_id == program.id)
            .filter(
                StudentSemester.status.notin_(
                    ["Deleted", "Deferred", "DroppedOut", "Withdrawn"]
                )
            )
            .order_by(StudentSemester.id)
            .all()
        )

        if not semesters:
            return None, "No Semesters Found"

        # Prepare semester data for CGPA calculation
        semesters_data = []
        for semester in semesters:
            # Get all student modules for this semester (excluding Delete/Drop status)
            modules = (
                db.query(StudentModule, SemesterModule.credits)
                .join(
                    SemesterModule,
                    StudentModule.semester_module_id == SemesterModule.id,
                )
                .filter(StudentModule.student_semester_id == semester.id)
                .filter(StudentModule.status.notin_(["Delete", "Drop"]))
                .all()
            )

            modules_data = []
            for student_module, credits in modules:
                grade = student_module.grade or ""

                modules_data.append(
                    {
                        "grade": grade,
                        "status": student_module.status,
                        "credits": float(credits),
                    }
                )

            semesters_data.append({"id": semester.id, "modules": modules_data})

        # Calculate CGPA using the comprehensive calculation
        grade_points, final_cgpa = calculate_cgpa_from_semesters(semesters_data)

        if final_cgpa == 0:
            return None, "No Valid Grades"

        # Round CGPA first to avoid floating-point precision issues
        # Classification should be based on the rounded value that students see
        rounded_cgpa = round(final_cgpa, 2)

        # Determine classification based on CGPA using grade descriptions
        if rounded_cgpa >= 3.5:  # A+, A, A- range (Pass with Distinction)
            classification = "Distinction"
        elif rounded_cgpa >= 3.0:  # B+, B, B- range (Pass with Merit)
            classification = "Merit"
        elif rounded_cgpa >= 1.7:  # C+, C, C- range (Pass)
            classification = "Pass"
        else:
            classification = "Failed"

        return rounded_cgpa, classification

    except Exception as e:
        click.echo(f"Error calculating CGPA for student {std_no}: {str(e)}")
        return None, "Calculation Error"


def calculate_cgpa_and_classification(
    db: Session, std_no: int
) -> Tuple[Optional[float], str]:
    """
    Calculate CGPA and determine classification for a student based on their active or latest completed program.
    Uses the same logic as the JavaScript implementation.

    Returns:
        Tuple of (CGPA, Classification)
    """
    try:
        # Get active program first, if not found get latest completed program
        active_program = (
            db.query(StudentProgram)
            .filter(
                and_(StudentProgram.std_no == std_no, StudentProgram.status == "Active")
            )
            .first()
        )

        if not active_program:
            # Try to get the latest completed program
            active_program = (
                db.query(StudentProgram)
                .filter(
                    and_(
                        StudentProgram.std_no == std_no,
                        StudentProgram.status == "Completed",
                    )
                )
                .order_by(StudentProgram.created_at.desc())
                .first()
            )

        if not active_program:
            return None, "No Active or Completed Program"

        # Get all semesters for the active program (excluding deleted/deferred/etc)
        semesters = (
            db.query(StudentSemester)
            .filter(StudentSemester.student_program_id == active_program.id)
            .filter(
                StudentSemester.status.notin_(
                    ["Deleted", "Deferred", "DroppedOut", "Withdrawn"]
                )
            )
            .order_by(StudentSemester.id)
            .all()
        )

        if not semesters:
            return None, "No Semesters Found"

        # Prepare semester data for CGPA calculation
        semesters_data = []
        for semester in semesters:
            # Get all student modules for this semester (excluding Delete/Drop status)
            modules = (
                db.query(StudentModule, SemesterModule.credits)
                .join(
                    SemesterModule,
                    StudentModule.semester_module_id == SemesterModule.id,
                )
                .filter(StudentModule.student_semester_id == semester.id)
                .filter(StudentModule.status.notin_(["Delete", "Drop"]))
                .all()
            )

            modules_data = []
            for student_module, credits in modules:
                grade = student_module.grade or ""

                modules_data.append(
                    {
                        "grade": grade,
                        "status": student_module.status,
                        "credits": float(credits),
                    }
                )

            semesters_data.append({"id": semester.id, "modules": modules_data})

        # Calculate CGPA using the comprehensive calculation
        grade_points, final_cgpa = calculate_cgpa_from_semesters(semesters_data)

        if final_cgpa == 0:
            return None, "No Valid Grades"

        # Round CGPA first to avoid floating-point precision issues
        # Classification should be based on the rounded value that students see
        rounded_cgpa = round(final_cgpa, 2)

        # Determine classification based on CGPA using grade descriptions
        if rounded_cgpa >= 3.5:  # A+, A, A- range (Pass with Distinction)
            classification = "Distinction"
        elif rounded_cgpa >= 3.0:  # B+, B, B- range (Pass with Merit)
            classification = "Merit"
        elif rounded_cgpa >= 1.7:  # C+, C, C- range (Pass)
            classification = "Pass"
        else:
            classification = "Failed"

        return rounded_cgpa, classification

    except Exception as e:
        click.echo(f"Error calculating CGPA for student {std_no}: {str(e)}")
        return None, "Calculation Error"


def get_student_classification(db: Session, std_no: int) -> Optional[str]:
    """
    Get the student's classification based on CGPA calculation.

    Deprecated: Use calculate_cgpa_and_classification instead.
    """
    _, classification = calculate_cgpa_and_classification(db, std_no)
    return classification


def export_graduating_students(
    db: Session,
    graduation_year: int,
    completion_terms: List[str],
    program_levels: List[str],
) -> None:
    """
    Export graduating students to Excel file.

    Graduating students are determined by:
    1. Students with approved academic graduation clearances (100% graduating), OR
    2. Students expected to graduate (based on reg_date or completion terms)
       who have NO pending academic issues

    Students expected to graduate but WITH pending issues go to "Non-Graduating Students" sheet.

    The exported file includes: student number, name, program name, school name, CGPA, classification, and criteria met.
    CGPA and classification are calculated from student's module grades in their active program.

    Args:
        db: Database session
        graduation_year: Year for graduation (e.g., 2025)
        completion_terms: List of completion terms to check (e.g., ["2025-02", "2024-07"])
        program_levels: List of program levels to include (e.g., ["diploma", "degree"])
    """
    graduating_students = []

    # Step 1: Get students with approved academic graduation clearances (100% graduating)
    # Filter by program levels
    click.echo("Finding students with approved academic graduation clearances...")

    approved_graduation_students = (
        db.query(StudentProgram.std_no)
        .join(
            GraduationRequest,
            StudentProgram.id == GraduationRequest.student_program_id,
        )
        .join(
            GraduationClearance,
            GraduationRequest.id == GraduationClearance.graduation_request_id,
        )
        .join(Clearance, GraduationClearance.clearance_id == Clearance.id)
        .join(Structure, StudentProgram.structure_id == Structure.id)
        .join(Program, Structure.program_id == Program.id)
        .filter(
            and_(
                Clearance.department == "academic",
                Clearance.status == "approved",
                Program.level.in_(program_levels),  # Filter by specified program levels
            )
        )
        .all()
    )

    approved_std_nos = {row.std_no for row in approved_graduation_students}
    click.echo(
        f"Found {len(approved_std_nos)} students with approved academic clearances (100% graduating)"
    )

    # Step 2: Get students expected to graduate
    expected_students = get_students_expected_to_graduate(
        db, graduation_year, completion_terms, program_levels
    )

    # Step 3: Separate expected students into graduating vs non-graduating based on pending issues
    click.echo("Filtering expected students by pending academic issues...")

    expected_graduating = []
    expected_with_issues = []

    for index, student in enumerate(expected_students, 1):
        if index % 50 == 0:
            click.echo(f"Checked {index}/{len(expected_students)} expected students...")

        std_no = student["std_no"]

        # Skip if already approved (they're 100% graduating)
        if std_no in approved_std_nos:
            continue

        # Check for pending issues
        if has_no_pending_issues(db, std_no):
            expected_graduating.append(student)
        else:
            expected_with_issues.append(student)

    click.echo(f"Expected students without pending issues: {len(expected_graduating)}")
    click.echo(f"Expected students with pending issues: {len(expected_with_issues)}")

    # Get non-graduating students (those with pending issues)
    non_graduating_students = get_non_graduating_students(db, expected_with_issues)

    # Step 4: Combine all graduating students
    all_graduating_std_nos = approved_std_nos | {
        s["std_no"] for s in expected_graduating
    }

    click.echo(f"Total graduating students: {len(all_graduating_std_nos)}")

    if not all_graduating_std_nos:
        click.secho("No graduating students found.", fg="yellow")
        return

    graduating_std_list = list(all_graduating_std_nos)

    # Get detailed information for all graduating students
    click.echo("Collecting student details...")

    # Preload student names for faster lookups
    student_names = (
        db.query(Student.std_no, Student.name)
        .filter(Student.std_no.in_(graduating_std_list))
        .all()
    )
    student_name_map = {row.std_no: row.name for row in student_names}

    # Prepare eager loading of related program data
    program_loader = (
        joinedload(StudentProgram.structure)
        .joinedload(Structure.program)
        .joinedload(Program.school)
    )

    # Preload active and completed programs for all candidates
    # For students with multiple completed programs, we'll select the latest one
    program_rows = (
        db.query(StudentProgram)
        .options(program_loader)
        .filter(
            and_(
                StudentProgram.std_no.in_(graduating_std_list),
                StudentProgram.status.in_(["Active", "Completed"]),
            )
        )
        .order_by(StudentProgram.std_no, StudentProgram.created_at.desc())
        .all()
    )

    # Build maps for active and latest completed programs
    active_program_map: Dict[int, StudentProgram] = {}
    completed_program_map: Dict[int, StudentProgram] = {}

    for program in program_rows:
        if program.status == "Active":
            active_program_map.setdefault(program.std_no, program)
        elif program.status == "Completed":
            # Only store the latest completed program (already sorted by created_at desc)
            completed_program_map.setdefault(program.std_no, program)

    # Preload approved programs using graduation requests
    approved_program_map: Dict[int, StudentProgram] = {}
    if approved_std_nos:
        approved_program_rows = (
            db.query(StudentProgram, GraduationRequest.created_at)
            .options(program_loader)
            .join(
                GraduationRequest,
                StudentProgram.id == GraduationRequest.student_program_id,
            )
            .join(
                GraduationClearance,
                GraduationRequest.id == GraduationClearance.graduation_request_id,
            )
            .join(Clearance, GraduationClearance.clearance_id == Clearance.id)
            .filter(
                and_(
                    StudentProgram.std_no.in_(graduating_std_list),
                    Clearance.department == "academic",
                    Clearance.status == "approved",
                )
            )
            .order_by(StudentProgram.std_no, GraduationRequest.created_at.desc())
            .all()
        )

        latest_request_per_student: Dict[int, int] = {}
        for program, created_at in approved_program_rows:
            created_at_value = created_at or 0
            std_no = program.std_no
            if (
                std_no not in latest_request_per_student
                or created_at_value > latest_request_per_student[std_no]
            ):
                approved_program_map[std_no] = program
                latest_request_per_student[std_no] = created_at_value

    for i, std_no in enumerate(graduating_std_list, 1):
        if i % 50 == 0:
            click.echo(f"Processed {i}/{len(all_graduating_std_nos)} students...")

        try:
            student_name = student_name_map.get(std_no)
            if not student_name:
                continue

            # Prefer approved program if available, otherwise prefer active over completed
            # For completed programs, use the latest one
            target_program = approved_program_map.get(std_no)
            if not target_program:
                target_program = active_program_map.get(std_no)
            if not target_program:
                target_program = completed_program_map.get(std_no)

            if not target_program:
                continue

            structure = target_program.structure
            program = structure.program if structure else None
            school = program.school if program and program.school else None

            program_name = program.name if program else "Unknown Program"
            school_name = school.name if school else "Unknown School"

            # Calculate CGPA using the target program (graduation request program or active program)
            cgpa, classification = calculate_cgpa_and_classification_for_program(
                db, std_no, target_program
            )

            # Determine graduation criteria met
            criteria_met = []
            if std_no in approved_std_nos:
                criteria_met.append("Approved Clearance")
            # Check if student is in expected graduating list
            for expected_student in expected_graduating:
                if expected_student["std_no"] == std_no:
                    criteria_met.append(expected_student["criteria"])
                    break

            graduating_students.append(
                {
                    "student_number": std_no,
                    "student_name": student_name,
                    "school_name": school_name,
                    "program_name": program_name,
                    "cgpa": cgpa if cgpa is not None else "N/A",
                    "classification": classification,
                    "criteria_met": " & ".join(criteria_met),
                }
            )

        except Exception as e:
            click.echo(f"Error processing student {std_no}: {str(e)}")
            continue

    if not graduating_students:
        click.secho("No valid graduating students found after processing.", fg="yellow")
        return

    # Filter out students with "Failed" classification
    initial_count = len(graduating_students)
    graduating_students = [
        s for s in graduating_students if s["classification"] != "Failed"
    ]
    failed_count = initial_count - len(graduating_students)

    if failed_count > 0:
        click.echo(f"Excluded {failed_count} students with 'Failed' classification")

    if not graduating_students:
        click.secho(
            "No graduating students remaining after filtering out failed classifications.",
            fg="yellow",
        )
        return

    # Calculate graduation statistics (we already have non_graduating_students from earlier)
    graduation_stats = calculate_graduation_statistics(
        db, graduating_students, non_graduating_students
    )

    # Sort students by school name, program name, then CGPA (descending)
    graduating_students.sort(
        key=lambda x: (
            x["school_name"],
            x["program_name"],
            -(
                x["cgpa"] if isinstance(x["cgpa"], (int, float)) else 0
            ),  # Negative for descending CGPA
        )
    )

    # Export to Excel
    output_dir = "exports"
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"graduating_students_{timestamp}.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)

    wb = Workbook()
    ws = wb.active
    if ws is None:
        # This should never happen with a new workbook, but let's handle it
        ws = wb.create_sheet("Graduating Students")
    else:
        ws.title = "Graduating Students"

    # Set up headers
    headers = [
        "Student Number",
        "Student Name",
        "School Name",
        "Program Name",
        "CGPA",
        "Classification",
        "Criteria Met",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(
        start_color="000000", end_color="000000", fill_type="solid"
    )

    header_widths = [len(header) for header in headers]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add data rows
    for row, student in enumerate(graduating_students, 2):
        row_values = [
            student["student_number"],
            student["student_name"],
            student["school_name"],
            student["program_name"],
            student["cgpa"],
            student["classification"],
            student["criteria_met"],
        ]
        for col_index, value in enumerate(row_values, 1):
            ws.cell(row=row, column=col_index, value=value)
            if value is not None:
                header_widths[col_index - 1] = max(
                    header_widths[col_index - 1], len(str(value))
                )

    for idx, width in enumerate(header_widths, 1):
        adjusted_width = min(width + 2, 50)
        ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    # Create breakdown sheet
    breakdown_ws = wb.create_sheet("School & Program Breakdown")

    # Prepare breakdown data
    from collections import defaultdict

    school_program_stats = defaultdict(lambda: defaultdict(int))
    school_totals = defaultdict(int)

    # Calculate statistics
    for student in graduating_students:
        school = student["school_name"]
        program = student["program_name"]
        school_program_stats[school][program] += 1
        school_totals[school] += 1

    # Set up breakdown sheet headers (use the same black header_fill)
    breakdown_headers = ["School/Faculty", "Program", "Student Count"]
    breakdown_widths = [len(header) for header in breakdown_headers]
    for col, header in enumerate(breakdown_headers, 1):
        cell = breakdown_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add breakdown data
    row = 2
    for school in sorted(school_program_stats.keys()):
        # Add school header row spanning three columns
        for col in range(1, 4):  # Columns 1, 2, 3
            if col == 1:
                cell = breakdown_ws.cell(row=row, column=col, value=school)
                breakdown_widths[0] = max(breakdown_widths[0], len(str(school)))
            else:
                cell = breakdown_ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="444444", end_color="444444", fill_type="solid"
            )
        row += 1

        # Add programs for this school
        programs = sorted(school_program_stats[school].keys())
        for program in programs:
            count = school_program_stats[school][program]
            breakdown_ws.cell(row=row, column=1, value="")  # Indent for program
            breakdown_ws.cell(row=row, column=2, value=program)
            breakdown_ws.cell(row=row, column=3, value=count)
            breakdown_widths[1] = max(breakdown_widths[1], len(str(program)))
            breakdown_widths[2] = max(breakdown_widths[2], len(str(count)))
            row += 1

        # Add school total
        total_cell = breakdown_ws.cell(row=row, column=2, value="Total")
        total_cell.font = Font(bold=True, color="444444")
        breakdown_widths[1] = max(breakdown_widths[1], len("Total"))

        total_count_cell = breakdown_ws.cell(
            row=row, column=3, value=school_totals[school]
        )
        total_count_cell.font = Font(bold=True, color="444444")
        breakdown_widths[2] = max(breakdown_widths[2], len(str(school_totals[school])))

        row += 2  # Add space between schools

    # Add grand total
    grand_total_row = row
    breakdown_ws.cell(row=grand_total_row, column=2, value="GRAND TOTAL")
    breakdown_ws.cell(row=grand_total_row, column=3, value=len(graduating_students))
    breakdown_widths[1] = max(breakdown_widths[1], len("GRAND TOTAL"))
    breakdown_widths[2] = max(breakdown_widths[2], len(str(len(graduating_students))))

    for col in [2, 3]:
        cell = breakdown_ws.cell(row=grand_total_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for idx, width in enumerate(breakdown_widths, 1):
        adjusted_width = min(width + 2, 60)
        breakdown_ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    # Create non-graduating students sheet
    if non_graduating_students:
        non_grad_ws = wb.create_sheet("Non-Graduating Students")

        # Set up headers for non-graduating students
        non_grad_headers = [
            "Student Number",
            "Student Name",
            "School Name",
            "Program Name",
            "Program Level",
            "Criteria",
            "Failed Never Repeated",
            "Never Attempted",
        ]

        non_grad_widths = [len(header) for header in non_grad_headers]
        for col, header in enumerate(non_grad_headers, 1):
            cell = non_grad_ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Add non-graduating student data
        for row, student in enumerate(non_graduating_students, 2):
            failed_modules = [
                f"{module['code']} - {module['name']}"
                for module in student["failed_never_repeated"]
            ]
            never_attempted_modules = [
                f"{module['code']} - {module['name']}"
                for module in student["never_attempted"]
            ]

            row_values = [
                student["student_number"],
                student["student_name"],
                student["school_name"],
                student["program_name"],
                student["program_level"].title(),
                student.get("criteria", ""),
                "; ".join(failed_modules) if failed_modules else "None",
                (
                    "; ".join(never_attempted_modules)
                    if never_attempted_modules
                    else "None"
                ),
            ]

            for col_index, value in enumerate(row_values, 1):
                non_grad_ws.cell(row=row, column=col_index, value=value)
                if value is not None:
                    non_grad_widths[col_index - 1] = max(
                        non_grad_widths[col_index - 1], len(str(value))
                    )

        for idx, width in enumerate(non_grad_widths, 1):
            adjusted_width = min(width + 2, 80)
            non_grad_ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    # Create graduation statistics sheet
    stats_ws = wb.create_sheet("Graduation Statistics")

    # Set up headers for statistics breakdown (use black header_fill)
    stats_headers = [
        "School/Faculty",
        "Program",
        "Expected",
        "Graduating",
        "Non-Graduating",
        "Graduation Rate (%)",
    ]

    for col, header in enumerate(stats_headers, 1):
        cell = stats_ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill

    # Add statistics breakdown data
    row = 2
    school_program_stats = graduation_stats["school_program_stats"]
    school_totals = graduation_stats["school_totals"]
    overall_stats = graduation_stats["overall_stats"]

    stats_widths = [len(header) for header in stats_headers]

    for school in sorted(school_program_stats.keys()):
        # Add school header row spanning all columns
        for col in range(1, 7):  # Columns 1-6
            if col == 1:
                cell = stats_ws.cell(row=row, column=col, value=school)
                stats_widths[0] = max(stats_widths[0], len(str(school)))
            else:
                cell = stats_ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(
                start_color="444444", end_color="444444", fill_type="solid"
            )
        row += 1

        # Add programs for this school
        programs = sorted(school_program_stats[school].keys())
        for program in programs:
            stats = school_program_stats[school][program]
            stats_ws.cell(row=row, column=1, value="")  # Indent for program
            stats_ws.cell(row=row, column=2, value=program)
            stats_ws.cell(row=row, column=3, value=stats["expected"])
            stats_ws.cell(row=row, column=4, value=stats["graduating"])
            stats_ws.cell(row=row, column=5, value=stats["non_graduating"])
            stats_ws.cell(row=row, column=6, value=f"{stats['percentage']:.1f}%")
            stats_widths[1] = max(stats_widths[1], len(str(program)))
            stats_widths[2] = max(stats_widths[2], len(str(stats["expected"])))
            stats_widths[3] = max(stats_widths[3], len(str(stats["graduating"])))
            stats_widths[4] = max(stats_widths[4], len(str(stats["non_graduating"])))
            stats_widths[5] = max(stats_widths[5], len(f"{stats['percentage']:.1f}%"))
            row += 1

        # Add school total
        school_stats = school_totals[school]
        total_cell = stats_ws.cell(row=row, column=2, value="Total")
        total_cell.font = Font(bold=True, color="444444")

        stats_ws.cell(row=row, column=3, value=school_stats["expected"])
        stats_ws.cell(row=row, column=4, value=school_stats["graduating"])
        stats_ws.cell(row=row, column=5, value=school_stats["non_graduating"])

        rate_cell = stats_ws.cell(
            row=row, column=6, value=f"{school_stats['percentage']:.1f}%"
        )
        rate_cell.font = Font(bold=True, color="444444")

        for col in [3, 4, 5]:
            cell = stats_ws.cell(row=row, column=col)
            cell.font = Font(bold=True, color="444444")

        stats_widths[1] = max(stats_widths[1], len("Total"))
        stats_widths[2] = max(stats_widths[2], len(str(school_stats["expected"])))
        stats_widths[3] = max(stats_widths[3], len(str(school_stats["graduating"])))
        stats_widths[4] = max(stats_widths[4], len(str(school_stats["non_graduating"])))
        stats_widths[5] = max(
            stats_widths[5], len(f"{school_stats['percentage']:.1f}%")
        )

        row += 2  # Add space between schools

    # Add grand total
    grand_total_row = row
    stats_ws.cell(row=grand_total_row, column=2, value="GRAND TOTAL")
    stats_ws.cell(row=grand_total_row, column=3, value=overall_stats["expected"])
    stats_ws.cell(row=grand_total_row, column=4, value=overall_stats["graduating"])
    stats_ws.cell(row=grand_total_row, column=5, value=overall_stats["non_graduating"])
    stats_ws.cell(
        row=grand_total_row, column=6, value=f"{overall_stats['percentage']:.1f}%"
    )

    stats_widths[1] = max(stats_widths[1], len("GRAND TOTAL"))
    stats_widths[2] = max(stats_widths[2], len(str(overall_stats["expected"])))
    stats_widths[3] = max(stats_widths[3], len(str(overall_stats["graduating"])))
    stats_widths[4] = max(stats_widths[4], len(str(overall_stats["non_graduating"])))
    stats_widths[5] = max(stats_widths[5], len(f"{overall_stats['percentage']:.1f}%"))

    for col in [2, 3, 4, 5, 6]:
        cell = stats_ws.cell(row=grand_total_row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    # Auto-size columns for statistics sheet
    for idx, width in enumerate(stats_widths, 1):
        adjusted_width = min(width + 2, 60)
        stats_ws.column_dimensions[get_column_letter(idx)].width = adjusted_width

    # Save the file
    wb.save(excel_path)

    click.secho(
        f"Successfully exported graduating students to: {excel_path}", fg="green"
    )

    # Display summary
    click.echo(f"\nSummary:")
    click.echo(
        f"- Total graduating students (excluding failed): {len(graduating_students)}"
    )
    if failed_count > 0:
        click.echo(
            f"- Students excluded due to 'Failed' classification: {failed_count}"
        )
    click.echo(f"- Students with approved academic clearances: {len(approved_std_nos)}")
    click.echo(
        f"- Expected students without pending issues: {len(expected_graduating)}"
    )
    click.echo(
        f"- Non-graduating students (expected but have pending issues): {len(non_graduating_students)}"
    )

    overall_stats = graduation_stats["overall_stats"]
    click.echo(
        f"- Overall graduation rate: {overall_stats['percentage']:.1f}% ({overall_stats['graduating']}/{overall_stats['expected']})"
    )

    # Show breakdown by school
    from collections import Counter

    school_counts = Counter(student["school_name"] for student in graduating_students)

    click.echo(f"\nSchool breakdown:")
    for school, count in school_counts.most_common():
        click.echo(f"- {school}: {count} students")

    # Show breakdown by program
    program_counts = Counter(student["program_name"] for student in graduating_students)

    click.echo(f"\nProgram breakdown:")
    for program, count in program_counts.most_common():
        click.echo(f"- {program}: {count} students")

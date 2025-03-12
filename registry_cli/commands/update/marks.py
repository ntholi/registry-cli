import os
from typing import cast

import click
import openpyxl
from sqlalchemy.orm import Session

from registry_cli.models import GradeType, StudentModule


def update_marks_from_excel(db: Session, file_path: str) -> None:

    if not os.path.exists(file_path):
        click.secho(f"Error: File {file_path} not found", fg="red")
        return

    try:
        click.echo(f"Reading data from {file_path}...")
        workbook = openpyxl.load_workbook(file_path, data_only=True)
        worksheet = workbook.active

        if worksheet is None:
            click.secho(f"Error: No active worksheet found in {file_path}", fg="red")
            return

        header_row = [cell.value for cell in worksheet[1]]
        required_columns = ["StdModuleID", "Final Mark", "Grade"]
        column_indices = {}
        missing_columns = []

        for column in required_columns:
            if column in header_row:
                column_indices[column] = header_row.index(column) + 1
            else:
                missing_columns.append(column)

        if missing_columns:
            click.secho(
                f"Error: Missing required columns: {', '.join(missing_columns)}",
                fg="red",
            )
            return

        status_column_index = None
        if "Status" in header_row:
            status_column_index = header_row.index("Status") + 1
        else:
            status_column_index = len(header_row) + 1
            worksheet.cell(row=1, column=status_column_index, value="Status")

        update_count = 0
        skipped_count = 0
        error_count = 0

        for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
            try:
                status_cell = worksheet.cell(row=row_index, column=status_column_index)
                if status_cell.value == "done":
                    skipped_count += 1
                    continue

                module_id_cell = row[column_indices["StdModuleID"] - 1]
                marks_cell = row[column_indices["Final Mark"] - 1]
                grade_cell = row[column_indices["Grade"] - 1]

                if module_id_cell.value is None:
                    continue

                try:
                    module_id = int(module_id_cell.value)
                except (ValueError, TypeError):
                    click.secho(
                        f"Warning: Invalid StdModuleID at row {row_index}: {module_id_cell.value}",
                        fg="yellow",
                    )
                    error_count += 1
                    continue

                marks = str(marks_cell.value) if marks_cell.value is not None else ""
                grade = str(grade_cell.value) if grade_cell.value is not None else ""

                module = (
                    db.query(StudentModule)
                    .filter(StudentModule.id == module_id)
                    .first()
                )

                if module:
                    module.marks = str(int(float(marks)))
                    try:
                        module.grade = cast(GradeType, grade)
                    except ValueError:
                        click.secho(
                            f"Warning: Invalid grade '{grade}' at row {row_index}, skipping grade update",
                            fg="yellow",
                        )
                        error_count += 1
                        continue

                    update_count += 1
                    worksheet.cell(
                        row=row_index, column=status_column_index, value="done"
                    )
                else:
                    click.secho(
                        f"Warning: No StudentModule found with ID {module_id}",
                        fg="yellow",
                    )
                    error_count += 1
            except Exception as e:
                click.secho(f"Error processing row {row_index}: {str(e)}", fg="red")
                error_count += 1

        db.commit()

        workbook.save(file_path)
        click.secho(f"Successfully updated {update_count} modules", fg="green")

        if skipped_count > 0:
            click.secho(
                f"Skipped {skipped_count} rows that were already marked as done",
                fg="blue",
            )

        if error_count > 0:
            click.secho(f"Encountered {error_count} errors while updating", fg="red")

    except Exception as e:
        db.rollback()
        click.secho(f"Error reading or processing Excel file: {str(e)}", fg="red")
